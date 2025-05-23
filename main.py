import os
import shutil
import pandas as pd
from sqlalchemy import create_engine, inspect
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import column_index_from_string
from PIL import Image as PILImage
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import logging
from datetime import datetime


# Variable global para almacenar el resultado de la consulta
resultado = None
fincas_vias_data = None

# Configurar logging
logging.basicConfig(filename='registro.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Definir el estilo del borde normal
normal_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

# Función para configurar la conexión a la base de datos
def obtener_conexion():
    conn_params = {
        'host': 'cartofs.seresco.red',
        'dbname': 'SEGIPSA',
        'user': 'Segipsa',
        'password': 'Segipas24',
        'port': 5432
    }
    conn_str = f"postgresql+psycopg2://{conn_params['user']}:{conn_params['password']}@{conn_params['host']}:{conn_params['port']}/{conn_params['dbname']}"
    return create_engine(conn_str)



# Función para listar las tablas en la base de datos bajo un esquema dado
def listar_tablas(esquema):
    engine = obtener_conexion()
    inspector = inspect(engine)
    tablas = inspector.get_table_names(schema=esquema)
    return tablas

# Función para obtener REFCAT de la base de datos
def obtener_refcat(esquema):
    engine = obtener_conexion()
    tablas = listar_tablas(esquema)
    logging.info(f"Tablas disponibles en el esquema '{esquema}': {tablas}")
    
    # Asegúrate de que la tabla 'segipsa_placo' está en la lista de tablas
    if 'segipsa_placo' not in tablas:
        raise ValueError(f"La tabla 'segipsa_placo' no existe en el esquema '{esquema}'. Tablas disponibles: " + ", ".join(tablas))
    
    # Completa la consulta con esquema.tabla
    query = f'SELECT "REFCAT" FROM "{esquema}"."segipsa_placo" WHERE "accion" != \'CIER\''
    logging.info("Consulta ejecutada: " + query)
    df = pd.read_sql_query(query, engine)
    logging.info(f"Consulta exitosa. Resultado: {df}")
    return df['REFCAT'].tolist()



def obtener_datos_por_refcat(esquema):
    global resultado
    engine = obtener_conexion()
    query = f"""
        SELECT exp, control, anio, del, mun, nom_mun, fecha_proyecto, fecha_licencia, fecha_act_urbanist, fecha_cert_finob, "REFCAT", sigla_via, situacion, npoli, dupli, cp, cod_incidencia, fecha_inf_ayunt, fecha_otras, cod_incidencia_adicional, tr_digi_grab, tr_campo, fecha_alt, justif_fecha_alteracion, observaciones
        FROM {esquema}.segipsa_placo
    """
    df = pd.read_sql_query(query, engine)
    resultado = df
    

    if df.empty:
        logging.warning(f"No se encontró el valor de 'observaciones'")
        return None
    
    return df

# Función para redimensionar y guardar las imágenes
def resize_image(image_path, output_path, width, height):
    with PILImage.open(image_path) as img:
        resized_img = img.resize((width, height), PILImage.Resampling.LANCZOS)
        resized_img.save(output_path)

# Función para añadir imágenes al archivo Excel
def add_images_to_excel(ws, png_path, jpg_path, esquema, refcat_value):
    global resultado
    global fincas_vias_data

    df = resultado

    # Iterar sobre cada fila del DataFrame
    for _, row in df.iterrows():
        if row['REFCAT'] == refcat_value:
            # Nombre del archivo usando el campo 'exp'
            exp_value = row['exp']
            file_name = f"FICHA_RESUMEN_PLACO23_{exp_value}.xlsx"
            file_path = os.path.join(esquema, file_name)

            logging.info(f"Expediente value: {exp_value}")

            # Procesar el campo_1 (agregar del y mun)
            del_value = str(row["del"]).zfill(2)
            mun_value = str(row["mun"]).zfill(3)
            nommun_value = row["nom_mun"]
            texto_agregar_1 = f"GERENCIA-MUNICIPIO: {del_value}{mun_value} {nommun_value}"
            ws["A7"].value = texto_agregar_1

            # Procesar el campo_2
            expediente = f"{row['exp']}.{row['control']}/{row['anio']}"
            texto_agregar_2 = f"Nº EXPEDIENTE: {expediente}"
            ws["P7"].value = texto_agregar_2

            # Procesar el campo_4 (REFCAT)
            REFCAT_value = row["REFCAT"]
            texto_agregar_4 = f"REFERENCIA CATASTRAL: {REFCAT_value}"
            ws["A9"].value = texto_agregar_4

            # Procesar el campo L7 (DIRECCIÓN) usando fincas_vias_data
            for finca_via in fincas_vias_data[1:]:  # Omitir la primera fila con las cabeceras
                # Verificar si cada campo es None y reemplazarlo por una cadena vacía si es necesario
                direccion_texto = (
                    f"DIRECCIÓN: "
                    f"{finca_via['SIGLA_DENOM'] if finca_via['SIGLA_DENOM'] is not None else ''} "
                    f"{finca_via['NUMERO'] if finca_via['NUMERO'] is not None else ''}"
                    f"{finca_via['NUMERO_DUP'] if finca_via['NUMERO_DUP'] is not None else ''} "
                    f"{row['cp'] if row['cp'] is not None else ''}"
                ).strip()

                ws["P9"].value = direccion_texto
                break

    # Insertar la imagen PNG si existe
    if os.path.exists(png_path):
        img_png = OpenpyxlImage(png_path)
        img_png.anchor = 'C13'  # Anclar la imagen en la celda C13
        ws.add_image(img_png)
    else:
        logging.warning(f"No se encontró la imagen PNG en la ruta: {png_path}")

    # Insertar la imagen JPG si existe
    if os.path.exists(jpg_path):
        img_jpg = OpenpyxlImage(jpg_path)
        img_jpg.anchor = 'P13'  # Anclar la imagen en la celda P13
        ws.add_image(img_jpg)
    else:
        logging.warning(f"No se encontró la imagen JPG en la ruta: {jpg_path}")


# Función para escribir datos iniciales en el Excel
def escribir_datos_iniciales(ws, esquema, refcat_value):
    global resultado
    global fincas_vias_data

    ##cabecera
    dfi = resultado

    # Ordenar DataFrame 'dfi' solo por 'ORD_CONS'
    if 'ORD_CONS' in dfi.columns:
        dfi = dfi.sort_values(by=['ORD_CONS'])
    else:
        logging.warning(f"La columna 'ORD_CONS' no está presente en el DataFrame")

    # Iterar sobre cada fila del DataFrame
    for _, row in dfi.iterrows():
        if row['REFCAT'] == refcat_value:
            # Nombre del archivo usando el campo 'exp'
            exp_value = row['exp']
            file_name = f"FICHA_RESUMEN_PLACO23_{exp_value}.xlsx"
            file_path = os.path.join(esquema, file_name)

            logging.info(f"Expediente value: {exp_value}")

            # Procesar el campo_1 (agregar del y mun)
            del_value = str(row["del"]).zfill(2)
            mun_value = str(row["mun"]).zfill(3)
            nommun_value = row["nom_mun"]
            texto_agregar_1 = f"GERENCIA-MUNICIPIO: {del_value}{mun_value} {nommun_value}"
            ws["A7"].value = texto_agregar_1

            # Procesar el campo_2
            expediente = f"{row['exp']}.{row['control']}/{row['anio']}"
            texto_agregar_2 = f"Nº EXPEDIENTE: {expediente}"
            ws["P7"].value = texto_agregar_2

            # Procesar el campo_4 (REFCAT)
            REFCAT_value = row["REFCAT"]
            texto_agregar_4 = f"REFERENCIA CATASTRAL: {REFCAT_value}"
            ws["A9"].value = texto_agregar_4

            # Procesar el campo L7 (DIRECCIÓN) usando fincas_vias_data
            for finca_via in fincas_vias_data[1:]:  # Omitir la primera fila con las cabeceras
                # Verificar si cada campo es None y reemplazarlo por una cadena vacía si es necesario
                direccion_texto = (
                    f"DIRECCIÓN: "
                    f"{finca_via['SIGLA_DENOM'] if finca_via['SIGLA_DENOM'] is not None else ''} "
                    f"{finca_via['NUMERO'] if finca_via['NUMERO'] is not None else ''}"
                    f"{finca_via['NUMERO_DUP'] if finca_via['NUMERO_DUP'] is not None else ''} "
                    f"{row['cp'] if row['cp'] is not None else ''}"
                ).strip()

                ws["P9"].value = direccion_texto
                break

    engine = obtener_conexion()
    
    mapeo_campos = {
        "CARGO": "A",
        "ORD_CONS": "B",
        "ES": "C",
        "PLA": "D",
        "PUE": "E",
        "TIPOL": "F",
        "CAT_PREDO": "G",
        "DES": "H",
        "SUP_LOCAL": "I",
        "U_CONS": "J",
        "AP_CO_CO": "K",
        "ANY_ANTIG": "L",
        "ANY_REFOR": "M"
    }

    start_row = 14
    query = f'SELECT * FROM "{esquema}"."DATOS_INICIALES" WHERE "REFCAT" = %(refcat)s'
    df_datos_iniciales = pd.read_sql_query(query, engine, params={'refcat': refcat_value})

    if df_datos_iniciales.empty:
        logging.error(f"No se encontraron registros con REFCAT = {refcat_value}")
        return False

    # Convertir los datos a una lista de tuplas para ordenarlos manualmente
    datos_ordenados = []
    for idx, row in df_datos_iniciales.iterrows():
        # Crear una tupla de (orden, fila completa)
        datos_ordenados.append((row['ORD_CONS'], row))

    # Ordenar la lista de tuplas por el primer elemento (ORD_CONS)
    datos_ordenados.sort(key=lambda x: x[0])

    # Escribir los datos ordenados en el Excel
    for idx, (ord_cons, row) in enumerate(datos_ordenados):
        for campo, col in mapeo_campos.items():
            if campo == "ANY_REFOR":
                # Solo escribir si ANY_REFOR es mayor que 0
                if row[campo] > 0:
                    valor_campo = row[campo]
                    dest_col = column_index_from_string(col)
                    dest_row = start_row + idx
                    ws.cell(row=dest_row, column=dest_col, value=valor_campo)
            else:
                if campo in row:
                    valor_campo = row[campo]
                    # Convertir el valor de CAT_PREDO a entero si es necesario
                    if campo == "CAT_PREDO" and not pd.isna(valor_campo):
                        valor_campo = int(valor_campo)
                    dest_col = column_index_from_string(col)
                    dest_row = start_row + idx
                    ws.cell(row=dest_row, column=dest_col, value=valor_campo)

        # Aplicar el estilo a toda la fila después de escribir los valores
        for col in range(1, ws.max_column + 1):
            ws.cell(row=dest_row, column=col)._style = ws.cell(row=start_row, column=col)._style
                
    return True





# Función para escribir datos de SAUCE en el Excel
def escribir_datos_sauce(ws, archivo_csv):
    secciones = {"FINCAS": [], "EXPEDIENTE": [], "CONSTRUCCIONES": [], "UNIDADES CONSTRUCTIVAS": [], "VIAS": []}
    seccion_actual = None

    with open(archivo_csv, 'r', encoding='utf-8') as file:
        for line in file:
            if line.strip() == "":
                seccion_actual = None
                continue

            if line.startswith("EXPEDIENTE"):
                seccion_actual = "EXPEDIENTE"
            elif line.startswith("FINCAS"):
                seccion_actual = "FINCAS"
            elif line.startswith("CONSTRUCCIONES"):
                seccion_actual = "CONSTRUCCIONES"
            elif line.startswith("UNIDADES CONSTRUCTIVAS"):
                seccion_actual = "UNIDADES CONSTRUCTIVAS"
            elif line.startswith("VIAS"):
                seccion_actual = "VIAS"
            elif seccion_actual:
                secciones[seccion_actual].append(line.strip())

    vias_data = {}
    for via in secciones["VIAS"]:
        campos_via = via.split(";")
        via_id = campos_via[0]
        vias_data[via_id] = {
            "SIGLA": campos_via[1],
            "DENOMINACION": campos_via[2]
        }

    fincas_data = []
    for finca in secciones["FINCAS"]:
        campos_finca = finca.split(";")
        finca_via = campos_finca[4]
        if finca_via in vias_data:#hago el cruce por codigo de via entre la seccion FINCAS y la sección VIAS
            sigla_denom = vias_data[finca_via]["SIGLA"] + " " + vias_data[finca_via]["DENOMINACION"]
            fincas_data.append({
                "VIA": finca_via,
                "NUMERO": campos_finca[5],
                "NUMERO_DUP": campos_finca[6],
                "SIGLA_DENOM": sigla_denom,
                "DENOMINACION": vias_data[finca_via]["DENOMINACION"]
            })

    construcciones_data = []
    for construccion in secciones["CONSTRUCCIONES"]:
        campos_construccion = construccion.split(";")
        construcciones_data.append([
            campos_construccion[29],  # Valor para "CARGO"
            campos_construccion[11],  # Valor para "ORDEN"
            campos_construccion[13],  # Valor para "ESCALERA"
            campos_construccion[14],  # Valor para "PLANTA"
            campos_construccion[15],  # Valor para "PUERTA"
            campos_construccion[19],  # Valor para "TIPOLOGIA"
            campos_construccion[20],  # Valor para "CATEG_PRED"
            campos_construccion[18],  # Valor para "DESTINO"
            campos_construccion[26],  # Valor para "SUPERFICIE_TOTAL"
            campos_construccion[16],  # Valor para "UNIDAD_CONST"
            "N/A",                    # Marcador de posición para "COEF_CONSERVACION"
            campos_construccion[23],  # Valor para "AA_ANTIGUEDAD"
            campos_construccion[22]   # Valor para "AA_REFORMA"
        ])
  

    for unidad in secciones["UNIDADES CONSTRUCTIVAS"]:
        campos_unidad = unidad.split(";")
        for construccion in construcciones_data:
            if construccion[9] == campos_unidad[4]:  # Verificar si "UNIDAD_CONST" coincide
                construccion[10] = campos_unidad[16]  # Actualizar "COEF_CONSERVACION"
                # No usar 'break' para seguir buscando más coincidencias

    # Construcciones_data tiene todas las filas de datos que necesitamos
    construcciones_data = construcciones_data[1:]
    
    mapeo_campos_sauce = {
        "CARGO": "O",
        "ORDEN": "P",
        "ESCALERA": "Q",
        "PLANTA": "R",
        "PUERTA": "S",
        "TIPOLOGIA": "T",
        "CATEG_PRED": "U",
        "DESTINO": "V",
        "SUPERFICIE_TOTAL": "W",
        "UNIDAD_CONST": "X",
        "COEF_CONSERVACION": "Y",
        "AA_ANTIGUEDAD": "Z",
        "AA_REFORMA": "AA"
    }
    
    start_row = 14
    for idx, construccion in enumerate(construcciones_data):
        for campo, col in mapeo_campos_sauce.items():
            valor_campo = construccion[list(mapeo_campos_sauce.keys()).index(campo)]
            ws[col + str(start_row + idx)] = valor_campo
            # Copiar formato de la fila anterior
            for col_num in range(1, ws.max_column + 1):
                ws.cell(row=start_row + idx, column=col_num)._style = ws.cell(row=start_row, column=col_num)._style

    # Variable global para almacenar la información combinada de FINCAS y VIAS
    global fincas_vias_data
    fincas_vias_data = []

    for finca in fincas_data:
        fincas_vias_data.append({
            "SIGLA_DENOM": finca["SIGLA_DENOM"],
            "DENOMINACION": finca["DENOMINACION"],
            "NUMERO": finca["NUMERO"],
            "NUMERO_DUP": finca["NUMERO_DUP"]
        })
            

# Función para comparar y resaltar diferencias
from openpyxl.styles import Border


# Función para comparar y resaltar diferencias
from openpyxl.styles import Font, Alignment, Border, Side

# Definición de bordes y alineación
thin_side = Side(border_style="thin", color="000000")
normal_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
center_alignment = Alignment(horizontal='center', vertical='center')

def comparar_y_resaltar(ws):
    # Definición de estilos
    thin_side = Side(border_style="thin", color="000000")
    normal_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Índices de columnas
    col_inicial_start = 1   # A
    col_inicial_end = 13    # M
    col_final_start = 15    # O
    col_final_end = 27      # AA

    row = 14
    max_row = ws.max_row

    while row <= max_row:
        # Valores de 'local' en inicial y final
        val_ini = str(ws.cell(row=row, column=2).value or "").strip()
        val_fin = str(ws.cell(row=row, column=16).value or "").strip()

        if val_ini and val_ini != val_fin:
            # Desplazar toda la columna O:AA hacia abajo a partir de esta fila
            rango = f"O{row}:AA{max_row}"
            ws.move_range(rango, rows=1, cols=0)
            max_row += 1
            # Estilo para la fila en blanco insertada
            for col in range(col_final_start, col_final_end + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = normal_border
                cell.alignment = center_alignment
            row += 1
            continue

        # Comparar e ir resaltando diferencias en texto
        for col in range(col_inicial_start, col_inicial_end + 1):
            cell_ini = ws.cell(row=row, column=col)
            cell_fin = ws.cell(row=row, column=col + (col_final_start - col_inicial_start))
            v_ini = str(cell_ini.value or "").strip()
            v_fin = str(cell_fin.value or "").strip()
            if v_ini != v_fin:
                cell_fin.font = Font(color="FF0000")

        row += 1

    # Aplicar bordes y alineación a todo el bloque inicial y final
    for block_start, block_end in [(col_inicial_start, col_inicial_end), (col_final_start, col_final_end)]:
        for r in ws.iter_rows(min_row=14, max_row=ws.max_row, min_col=block_start, max_col=block_end):
            for cell in r:
                cell.border = normal_border
                cell.alignment = center_alignment

    # Alinear resto de celdas (encabezados, etc.)
    for r in ws.iter_rows():
        for cell in r:
            cell.alignment = center_alignment



    
# Función para escribir en la ficha resumen
def escribir_ficha_resumen(ws, esquema, refcat_value):
    global resultado
    global fincas_vias_data
    #cabecera
    df = resultado
    
    # Iterar sobre cada fila del DataFrame
    for _, row in df.iterrows():
        if(row['REFCAT'] == refcat_value):
            # Nombre del archivo usando el campo 'exp'
            exp_value = row['exp']
            file_name = f"FICHA_RESUMEN_PLACO23_{exp_value}.xlsx"
            file_path = os.path.join(esquema, file_name)

            logging.info(f"Expediente value: {exp_value}")

            # Procesar el campo_1 (agregar del y mun)
            del_value = str(row["del"]).zfill(2)
            mun_value = str(row["mun"]).zfill(3)
            nommun_value = row["nom_mun"]
            texto_agregar_1 = f"GERENCIA-MUNICIPIO: {del_value}{mun_value} {nommun_value}"
            ws["A7"].value = texto_agregar_1

            # Procesar el campo_2
            expediente = f"{row['exp']}.{row['control']}/{row['anio']}"
            texto_agregar_2 = f"Nº EXPEDIENTE: {expediente}"
            ws["P7"].value = texto_agregar_2

            # Procesar las fechas en los campos correspondientes
            def procesar_fecha(celda, fecha, celda_fecha):
                if pd.notnull(fecha) and fecha != "":
                    # Convertir la fecha a string y luego al nuevo formato
                    fecha_obj = datetime.strptime(str(fecha), '%Y-%m-%d')
                    fecha_formateada = fecha_obj.strftime('%d/%m/%Y')
                    ws[celda].value = "x"
                    ws[celda_fecha].value = f"FECHA: {fecha_formateada}"
                else:
                    ws[celda_fecha].value = f"FECHA:"

            procesar_fecha('G20', row['fecha_proyecto'], 'H21')
            procesar_fecha('O20', row['fecha_licencia'], 'P21')
            procesar_fecha('V20', row['fecha_act_urbanist'], 'W21')
            procesar_fecha('G22', row['fecha_cert_finob'], 'H23')
            procesar_fecha('O22', row['fecha_inf_ayunt'], 'P23')
            procesar_fecha('V22', row['fecha_otras'], 'W23')

            # Procesar el campo_4 (REFCAT)
            REFCAT_value = row["REFCAT"]
            texto_agregar_4 = f"REFERENCIA CATASTRAL: {REFCAT_value}"
            ws["A9"].value = texto_agregar_4

            # Procesar el campo L7 (DIRECCIÓN) usando fincas_vias_data
            for finca_via in fincas_vias_data[1:]:  # Omitir la primera fila con las cabeceras
                # Verificar si cada campo es None y reemplazarlo por una cadena vacía si es necesario
                direccion_texto = (
                    f"DIRECCIÓN: "
                    f"{finca_via['SIGLA_DENOM'] if finca_via['SIGLA_DENOM'] is not None else ''} "
                    f"{finca_via['NUMERO'] if finca_via['NUMERO'] is not None else ''}"
                    f"{finca_via['NUMERO_DUP'] if finca_via['NUMERO_DUP'] is not None else ''} "
                    f"{row['cp'] if row['cp'] is not None else ''}"
                ).strip()

                ws["P9"].value = direccion_texto
                break

            # Procesar el campo_5 (cod_incidencia y cod_incidencia_adicional)
            # Diccionario con códigos y descripciones de tipos de incidencias
            tipos_incidencias = {
                "CPAR": "Contorno de parcela o mal cartografiado",
                "DME": "Parcela o construcción mal geo-referenciada",
                "BALSA": "Omisión de Balsa de riego",
                "PCONI": "Omisión de la explotación",
                "PINFR": "Omisión de infraestructura puntual",
                "CCUL": "Cambio de cultivo o aprovechamiento",
                "CUSO": "Cambio de uso",
                "TIPO": "Tipología o Categoría errónea",
                "NULO": "Código para indicar que no hay que poner incidencia aquí",
                "DOCC": "Construcción rústica catastrada pero no valorada",
                "DEXT": "Declaración extemporánea",
                "CIC": "Ámbito desactualizado",
                "DMA": "Ámbito mal geo-referenciado",
                "PINFL": "Omisión de infraestructura lineal",
                "DESU": "Omisión de desarrollo urbanístico",
                "HSOL": "Omisión de huerto solar",
                "PPEOL": "Omisión de parque eólico",
                "RECC": "Es necesario hacer recorrido de campo",
                "SDES": "DEMOLICION TOTAL",
                "NDES": "DEMOLICON PARCIAL",
                "NCON": "Ampliación u omisión de alguna construcción",
                "NPIS": "Omisión de piscina",
                "PCON": "Alta de obra nueva",
                "REHAB": "Rehabilitación",
                "REFOR": "Reforma",
                "ACT": "Actualización del estado de conservación"
            }

            # Procesar el campo_5 (cod_incidencia y cod_incidencia_adicional)
            cod_incidencia = row["cod_incidencia"]
            cod_incidencia_adicional = row["cod_incidencia_adicional"]

            if pd.notnull(cod_incidencia) and cod_incidencia in tipos_incidencias:
                descripcion_principal = tipos_incidencias[cod_incidencia].upper()
            else:
                descripcion_principal = cod_incidencia  # En caso de que el código no esté en el diccionario

            if pd.notnull(cod_incidencia_adicional) and cod_incidencia_adicional in tipos_incidencias:
                descripcion_adicional = tipos_incidencias[cod_incidencia_adicional].upper()
                texto_agregar_5 = f"{descripcion_principal}/{descripcion_adicional}"
            else:
                texto_agregar_5 = descripcion_principal

            ws['G11'].value = texto_agregar_5

            # Obtener el valor del campo Tr_digi_grab de tus datos
            tr_digi_grab_value = row["tr_digi_grab"]
            # Establecer el valor del campo en la celda K13
            ws['K13'] = tr_digi_grab_value
            # Establecer el estilo de fuente en cursiva para la celda K13
            ws['K13'].font = Font(italic=True)

            # Obtener el valor del campo Tr_campo de tus datos
            tr_campo_value = row["tr_campo"]
            # Establecer el valor del campo en la celda K16
            ws['K16'] = tr_campo_value
            # Establecer el estilo de fuente en cursiva para la celda K16
            ws['K16'].font = Font(italic=True)

            # Procesar el campo fecha_alt en la celda G25
            fecha_alt_value = row["fecha_alt"]
            if pd.notnull(fecha_alt_value) and fecha_alt_value != "":
                # Convertir la fecha a string y luego al nuevo formato
                fecha_alt_obj = datetime.strptime(str(fecha_alt_value), '%Y-%m-%d')
                fecha_alt_formateada = fecha_alt_obj.strftime('%d/%m/%Y')
                texto_agregar_fecha_alt = f"FECHA: {fecha_alt_formateada}"
            else:
                texto_agregar_fecha_alt = "FECHA:"
            # Establecer el estilo de la fuente en negrita para la celda G25
            ws['G25'].font = Font(bold=True)
            # Asignar el valor a la celda G25
            ws['G25'].value = texto_agregar_fecha_alt

            # Procesar el campo justif_fecha_alteracion en la celda G26
            justif_fecha_alteracion_value = row["justif_fecha_alteracion"]
            texto_agregar_justif_fecha_alteracion = f"MOTIVACIÓN: {justif_fecha_alteracion_value}"
            # Asignar el valor a la celda G26
            ws['G26'].value = texto_agregar_justif_fecha_alteracion

            # Procesar el campo de observaciones en la celda G29
            observaciones_value = row["observaciones"]
            # Asignar el valor a la celda G29
            ws['G29'].value = observaciones_value

# Función para procesar cada carpeta en el directorio de origen
def process_folders(window, output_dir, template_file, origin_dir, esquema, progress_label, progress_bar):
    refcat_list = obtener_refcat(esquema)
    total_folders = len(refcat_list)

    obtener_datos_por_refcat(esquema)

    for index, refcat_value in enumerate(refcat_list):
        refcat_value = refcat_value.strip()  # Asegúrate de que no haya espacios en blanco
        folder_path = os.path.join(origin_dir, refcat_value)  # Modificado para usar la ruta de origen
        excel_path = os.path.join(output_dir, f"{refcat_value}_FichaResumen.xlsx")
        png_path = os.path.join(folder_path, f"{refcat_value}.png")
        jpg_path = os.path.join(folder_path, f"{refcat_value}.jpg")
        csv_path = os.path.join(folder_path, f"{refcat_value}.csv")
        
        logging.info(f"Refcat value: {refcat_value}")
        
        # Actualizar la barra de progreso
        progress_bar['value'] = (index + 1) / total_folders * 100
        progress_label.config(text=f"Procesando carpeta {index + 1} de {total_folders}")
        window.update_idletasks()
        
        # Verificar si los archivos PNG, JPG y CSV existen en la carpeta actual
        if not all(os.path.exists(path) for path in [png_path, jpg_path, csv_path]):
            logging.warning(f"Archivos faltantes para REFCAT {refcat_value}")
            missing_files = [path for path in [png_path, jpg_path, csv_path] if not os.path.exists(path)]
            for missing_file in missing_files:
                logging.warning(f"El archivo {missing_file} no se encuentra en la carpeta {folder_path}")
            continue
        
        # Verificar si la carpeta de salida existe
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)  # Asegurarse de que la carpeta de salida existe
            logging.info(f"Carpeta de salida creada en: {output_dir}")
        
        # Copiar la plantilla al directorio de salida
        shutil.copy(template_file, excel_path)
        logging.info(f"Plantilla copiada a: {excel_path}")
        
        # Cargar el libro de trabajo y las hojas necesarias
        wb = openpyxl.load_workbook(excel_path)
        ws_iniciales = wb['SAUCE']
        ws_sauce = wb['SAUCE']
        ws_croquis = wb['CROQUIS']
        ws_ficha_resumen_placo = wb['FICHA RESUMEN PLACO']
        
        # Escribir datos de SAUCE en el libro de trabajo
        escribir_datos_sauce(ws_sauce, csv_path)
      
        # Agregar imágenes al libro de trabajo
        add_images_to_excel(ws_croquis, png_path, jpg_path, esquema, refcat_value)

      # Escribir datos iniciales en el libro de trabajo
        if not escribir_datos_iniciales(ws_iniciales, esquema, refcat_value):
            continue
        
        # Comparar y resaltar diferencias en los datos de SAUCE
        comparar_y_resaltar(ws_sauce)
        
        # Escribir datos de ficha resumen en el libro de trabajo
        escribir_ficha_resumen(ws_ficha_resumen_placo, esquema, refcat_value)
       
        
        # Guardar el libro de trabajo modificado
        wb.save(excel_path)
        logging.info(f"Libro de trabajo guardado en: {excel_path}")


# Función para abrir el explorador de archivos
def seleccionar_directorio(titulo):
    root = tk.Tk()
    root.withdraw()
    directorio = filedialog.askdirectory(title=titulo)
    return directorio

def seleccionar_archivo(titulo):
    root = tk.Tk()
    root.withdraw()
    archivo = filedialog.askopenfilename(title=titulo, filetypes=[('Excel Files', '*.xlsx')])
    return archivo

# Función para la interfaz gráfica
def interfaz_grafica():
    window = tk.Tk()
    window.title("Procesar Archivos Masivamente")

    ttk.Label(window, text="Esquema:").grid(column=0, row=0)
    schema_entry = ttk.Entry(window, width=50)
    schema_entry.grid(column=1, row=0)
    
    ttk.Label(window, text="Directorio de Salida:").grid(column=0, row=1)
    output_dir_entry = ttk.Entry(window, width=50)
    output_dir_entry.grid(column=1, row=1)
    ttk.Button(window, text="Seleccionar", command=lambda: output_dir_entry.insert(0, seleccionar_directorio("Seleccionar Directorio de Salida"))).grid(column=2, row=1)
    
    ttk.Label(window, text="Archivo de Plantilla:").grid(column=0, row=2)
    template_file_entry = ttk.Entry(window, width=50)
    template_file_entry.grid(column=1, row=2)
    ttk.Button(window, text="Seleccionar", command=lambda: template_file_entry.insert(0, seleccionar_archivo("Seleccionar Archivo de Plantilla"))).grid(column=2, row=2)
    
    ttk.Label(window, text="Directorio de Origen de Datos:").grid(column=0, row=3)
    origin_dir_entry = ttk.Entry(window, width=50)
    origin_dir_entry.grid(column=1, row=3)
    ttk.Button(window, text="Seleccionar", command=lambda: origin_dir_entry.insert(0, seleccionar_directorio("Seleccionar Directorio de Origen de Datos"))).grid(column=2, row=3)
    
    progress_label = ttk.Label(window, text="")
    progress_label.grid(column=0, row=4, columnspan=3)

    progress_bar = ttk.Progressbar(window, orient=tk.HORIZONTAL, length=300, mode='determinate')
    progress_bar.grid(column=0, row=5, columnspan=3)
        
    def ejecutar_proceso():
        esquema = schema_entry.get()
        output_dir = output_dir_entry.get()
        template_file = template_file_entry.get()
        origin_dir = origin_dir_entry.get()
        
        if not all([esquema, output_dir, template_file, origin_dir]):
            messagebox.showerror("Error", "Todos los campos son obligatorios.")
            return
        
        process_folders(window, output_dir, template_file, origin_dir, esquema, progress_label, progress_bar)
        messagebox.showinfo("Éxito", "Proceso completado.")

    ttk.Button(window, text="Ejecutar", command=ejecutar_proceso).grid(column=0, row=6, columnspan=3)
    
    window.mainloop()

# Ejecutar la interfaz gráfica
if __name__ == "__main__":
    interfaz_grafica()
