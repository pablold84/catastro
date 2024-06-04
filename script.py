import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox
import logging

# Configurar el logger
logging.basicConfig(filename='registro.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Función para configurar la conexión a la base de datos
def obtener_conexion():
    conn_params = {
        'host': 'cartofs.seresco.red',
        'dbname': 'SEGIPSA',
        'user': 'Segipsa',
        'password': 'Segipas24',
        'port': 5432
    }
    conn_str = f"postgresql+psycopg2://{conn_params['user']}:{conn_params['password']}@{conn_params['host']}:{conn_params['port']}/{conn_params['dbname']}"
    return create_engine(conn_str)

# Función para ejecutar el proceso
def ejecutar_proceso(output_dir, template_file):
    try:
        # Agregar registro de inicio del proceso
        logging.info('Iniciando el proceso de generación de archivos Excel.')
        
        # Conectar a la base de datos
        engine = obtener_conexion()
        
        # Obtener todos los registros de la base de datos
        query = """
            SELECT exp, control, anio, del, mun, nom_mun, fecha_proyecto, fecha_licencia, fecha_act_urbanist, fecha_cert_finob,   "REFCAT", sigla_via, situacion, npoli, dupli, cp, cod_incidencia, fecha_inf_ayunt, fecha_otras, cod_incidencia_adicional, tr_digi_grab, tr_campo, fecha_alt, justif_fecha_alteracion, observaciones
            FROM cabrales.segipsa_placo
            """
        df = pd.read_sql_query(query, engine)
        
        # Crear el directorio de salida si no existe
        os.makedirs(output_dir, exist_ok=True)
        
        # Iterar sobre cada registro y generar un archivo Excel
        for idx, row in df.iterrows():
            # Nombre del archivo usando el campo 'exp'
            exp_value = row['exp']
            file_name = f"FICHA_RESUMEN_PLACO23_{exp_value}.xlsx"
            file_path = os.path.join(output_dir, file_name)

            # Copiar el archivo de plantilla
            shutil.copyfile(template_file, file_path)

            # Cargar el archivo copiado
            book = load_workbook(file_path)
            ws = book["FICHA RESUMEN PLACO"]

            # Procesar el campo_1 (agregar del y mun)
            del_value = str(row["del"]).zfill(2)
            mun_value = str(row["mun"]).zfill(3)
            nommun_value = row["nom_mun"]
            texto_agregar_1 = f"GERENCIA-MUNICIPIO: {del_value}{mun_value} {nommun_value}"
            ws["A7"].value = texto_agregar_1

            # Procesar el campo_2
            expediente = f"{row['exp']}.{row['control']}/{row['anio']}"
            texto_agregar_2 = f"Nº EXPEDIENTE: {expediente}"
            ws["P7"].value = texto_agregar_2

            # Procesar el campo_3
            fecha_proyecto = row['fecha_proyecto']
            if pd.notnull(fecha_proyecto) and fecha_proyecto != "":
                fecha_proyecto_str = str(fecha_proyecto)
                ws['G20'].value = "x"
                ws['H21'].value = f"FECHA: {fecha_proyecto_str}"
            else:
                ws['H21'].value = f"FECHA:"

            fecha_licencia = row['fecha_licencia']
            if pd.notnull(fecha_proyecto) and fecha_licencia != "":
                fecha_licencia_str = str(fecha_licencia)
                ws['O20'].value = "x"
                ws['P21'].value = f"FECHA: {fecha_licencia_str}"
            else:
                ws['P21'].value = f"FECHA:"

            fecha_act_urbanist = row['fecha_act_urbanist']
            if pd.notnull(fecha_act_urbanist) and fecha_act_urbanist != "":
                fecha_act_urbanist_str = str(fecha_act_urbanist)
                ws['V20'].value = "x"
                ws['W21'].value = f"FECHA: {fecha_act_urbanist_str}"
            else:
                ws['W21'].value = f"FECHA:"

            fecha_cert_finob = row['fecha_cert_finob']
            if pd.notnull(fecha_cert_finob) and fecha_cert_finob != "":
                fecha_cert_finob_str = str(fecha_cert_finob)
                ws['G22'].value = "x"
                ws['H23'].value = f"FECHA: {fecha_cert_finob_str}"
            else:
                ws['H23'].value = f"FECHA:"

            fecha_inf_ayunt = row['fecha_inf_ayunt']
            if pd.notnull(fecha_inf_ayunt) and fecha_inf_ayunt != "":
                fecha_inf_ayunt_str = str(fecha_inf_ayunt)
                ws['O22'].value = "x"
                ws['P23'].value = f"FECHA: {fecha_inf_ayunt_str}"
            else:
                ws['P23'].value = f"FECHA:"

            fecha_otras = row['fecha_otras']
            if pd.notnull(fecha_otras) and fecha_otras != "":
                fecha_otras_str = str(fecha_otras)
                ws['V22'].value = "x"
                ws['W23'].value = f"FECHA: {fecha_otras_str}"
            else:
                ws['W23'].value = f"FECHA:"

            # Procesar el campo_4 (REFCAT)
            REFCAT_value = row["REFCAT"]
            texto_agregar_4 = f"REFERENCIA CATASTRAL: {REFCAT_value}"
            ws["A9"].value = texto_agregar_4

         
            # Procesar el campo L7 (DIRECCIÓN)
            direccion_texto = f"DIRECCIÓN: {row['sigla_via']} {row['situacion']} {row['npoli']}{row['dupli']} {row['cp']}"
            ws["P9"].value = direccion_texto

            # Procesar el campo_5 (cod_incidencia y cod_incidencia_adicional)
            cod_incidencia = row["cod_incidencia"]
            cod_incidencia_adicional = row["cod_incidencia_adicional"]
            if pd.notnull(cod_incidencia_adicional) and cod_incidencia_adicional != "":
                texto_agregar_5 = f"{cod_incidencia}/{cod_incidencia_adicional}"
            else:
                texto_agregar_5 = str(cod_incidencia)
            ws['G11'].value = texto_agregar_5

            # Obtener el valor del campo Tr_digi_grab de tus datos
            tr_digi_grab_value = row["tr_digi_grab"]
            # Establecer el valor del campo en la celda K16
            ws['K13'] = tr_digi_grab_value
            # Establecer el estilo de fuente en cursiva para la celda K16
            ws['K13'].font = Font(italic=True)

            # Obtener el valor del campo Tr_campo de tus datos
            tr_campo_value = row["tr_campo"]
            # Establecer el valor del campo en la celda K16
            ws['K16'] = tr_campo_value
            # Establecer el estilo de fuente en cursiva para la celda K16
            ws['K16'].font = Font(italic=True)

            # Procesar el campo fecha_alt en la celda G25
            fecha_alt_value = row["fecha_alt"]
            texto_agregar_fecha_alt = f"FECHA: {fecha_alt_value}"
            # Establecer el estilo de la fuente en negrita para la celda G25
            ws['G25'].font = Font(bold=True)
            # Asignar el valor a la celda G25
            ws['G25'].value = texto_agregar_fecha_alt

            # Procesar el campo justif_fecha_alteracion en la celda G26
            justif_fecha_alteracion_value = row["justif_fecha_alteracion"]
            texto_agregar_justif_fecha_alteracion = f"MOTIVACIÓN: {justif_fecha_alteracion_value}"
            # Asignar el valor a la celda G26
            ws['G26'].value = texto_agregar_justif_fecha_alteracion

            # Procesar el campo de observaciones en la celda G29
            observaciones_value = row["observaciones"]
            # Asignar el valor a la celda G29
            ws['G29'].value = observaciones_value

            # Guardar el archivo modificado
            book.save(file_path)
            logging.info(f"Archivo guardado: {file_path}")  # Mensaje de depuración

        messagebox.showinfo("Éxito", "Datos escritos exitosamente en los archivos Excel generados")
        logging.info('Proceso completado exitosamente.')

    except Exception as e:
        messagebox.showerror("Error", str(e))
        logging.error(f"Error durante el proceso: {str(e)}")

# Función para seleccionar el directorio de salida
def seleccionar_directorio():
    directorio = filedialog.askdirectory()
    if directorio:
        entry_output_dir.delete(0, tk.END)
        entry_output_dir.insert(0, directorio)

# Función para seleccionar el archivo de plantilla
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if archivo:
        entry_template_file.delete(0, tk.END)
        entry_template_file.insert(0, archivo)

# Crear la ventana principal
root = tk.Tk()
root.title("Generador de Modelo 6")

# Crear los widgets
label_output_dir = tk.Label(root, text="Directorio de Salida:")
label_output_dir.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
entry_output_dir = tk.Entry(root, width=50)
entry_output_dir.grid(row=0, column=1, padx=10, pady=5)
button_output_dir = tk.Button(root, text="Seleccionar", command=seleccionar_directorio)
button_output_dir.grid(row=0, column=2, padx=10, pady=5)

label_template_file = tk.Label(root, text="Archivo de Plantilla:")
label_template_file.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
entry_template_file = tk.Entry(root, width=50)
entry_template_file.grid(row=1, column=1, padx=10, pady=5)
button_template_file = tk.Button(root, text="Seleccionar", command=seleccionar_archivo)
button_template_file.grid(row=1, column=2, padx=10, pady=5)

button_ejecutar = tk.Button(root, text="Ejecutar", command=lambda: ejecutar_proceso(entry_output_dir.get(), entry_template_file.get()))
button_ejecutar.grid(row=2, column=1, padx=10, pady=20)

# Ejecutar la ventana principal
root.mainloop()

