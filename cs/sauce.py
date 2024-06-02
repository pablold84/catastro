import os
import shutil
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog

# Carga de la info SITUACIÓN FINAL datos sauce

# Definir el nombre del archivo CSV
archivo_csv = "modelo6/origen/ficheros/sauce/33_217210_24.csv"

# Diccionario para almacenar las líneas de cada sección
secciones = {"FINCAS": [], "EXPEDIENTE": [], "CONSTRUCCIONES": [], "UNIDADES CONSTRUCTIVAS": []}

# Bandera para indicar la sección actual
seccion_actual = None

# Leer el archivo CSV
with open(archivo_csv, 'r') as file:
    for line in file:
        # Verificar si la línea está vacía (contiene solo el retorno de carro)
        if line.strip() == "":
            # Si la línea está vacía, cambiar a la siguiente sección
            seccion_actual = None
            continue

        # Detectar la sección actual
        if line.startswith("EXPEDIENTE"):
            seccion_actual = "EXPEDIENTE"
            continue
        elif line.startswith("FINCAS"):
            seccion_actual = "FINCAS"
            continue
        elif line.startswith("CONSTRUCCIONES"):
            seccion_actual = "CONSTRUCCIONES"
            continue
        elif line.startswith("UNIDADES CONSTRUCTIVAS"):
            seccion_actual = "UNIDADES CONSTRUCTIVAS"
            continue

        # Almacenar las líneas en la sección actual si no está vacía
        if seccion_actual in secciones and line.strip() != "":
            secciones[seccion_actual].append(line.strip())

# Extraer datos de CONSTRUCCIONES
construcciones_data = []
for construccion in secciones["CONSTRUCCIONES"]:
    campos_construccion = construccion.split(";")
    cargo = campos_construccion[29]
    orden = campos_construccion[11]

    pcatastral1 = campos_construccion[9]
    pcatastral2 = campos_construccion[10]

    escalera = campos_construccion[13]
    planta = campos_construccion[14]
    puerta = campos_construccion[15]
    tipologia = campos_construccion[19]
    categ_pred = campos_construccion[20]
    destino = campos_construccion[18]
    superficie_total = campos_construccion[26]
    unidad_const_dest = campos_construccion[16]
    aa_antiguedad = campos_construccion[23]
    aa_reforma = campos_construccion[22]

    # Obtener el coeficiente de conservación de UNIDADES CONSTRUCTIVAS
    coef_conservacion = "N/A"
    for unidad in secciones["UNIDADES CONSTRUCTIVAS"]:
        campos_unidad = unidad.split(";")  # asigno si coincide refCat y unidad constructiva
        if pcatastral1 == campos_unidad[2] and pcatastral2 == campos_unidad[3] and unidad_const_dest == campos_unidad[4]:
            coef_conservacion = campos_unidad[16]
            break

    construcciones_data.append([cargo, orden, escalera, planta, puerta, tipologia, categ_pred, destino, superficie_total, unidad_const_dest, coef_conservacion, aa_antiguedad, aa_reforma])

# Omitir el primer elemento (encabezado) de construcciones_data
construcciones_data = construcciones_data[1:]

# Definir mapeo_campos_sauce
mapeo_campos_sauce = {
    "CARGO": "O",
    "ORDEN": "P",
    "ESCALERA": "Q",
    "PLANTA": "R",
    "PUERTA": "S",
    "TIPOLOGIA": "T",
    "CATEG_PRED": "U",
    "DESTINO": "V",
    "SUPERFICIE_TOTAL": "W",
    "UNIDAD_CONST": "X",
    "COEF_CONSERVACION": "Y",
    "AA_ANTIGUEDAD": "Z",
    "AA_REFORMA": "AA"
}

# Función para escribir los datos en un archivo Excel
def escribir_datos_en_archivo(template_file, output_dir, mapeo_campos_sauce, construcciones_data):
    # Crear el directorio de salida si no existe
    os.makedirs(output_dir, exist_ok=True)

    # Copiar el archivo de plantilla al directorio de salida
    file_name = os.path.basename(template_file)
    file_path = os.path.join(output_dir, file_name)
    shutil.copyfile(template_file, file_path)

    # Cargar el archivo de Excel
    book = load_workbook(file_path)
    ws = book["SAUCE"]

    print(construcciones_data)

    # Escribir los datos en el archivo
    start_row = 5  # Corresponde a la fila 5 en Excel
    for idx, construccion in enumerate(construcciones_data):
        for campo, col in mapeo_campos_sauce.items():
            valor_campo = construccion[list(mapeo_campos_sauce.keys()).index(campo)]
            ws[col + str(start_row + idx)] = valor_campo

    # Guardar el libro de Excel
    book.save(file_path)

    print("Datos escritos exitosamente en el archivo Excel generado")

# Crear la ventana principal
root = tk.Tk()
root.title("Escritor de Datos en Archivo Excel")

# Función para seleccionar el archivo de plantilla
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if archivo:
        entry_template_file.delete(0, tk.END)
        entry_template_file.insert(0, archivo)

# Función para seleccionar el directorio de salida
def seleccionar_directorio():
    directorio = filedialog.askdirectory()
    if directorio:
        entry_output_dir.delete(0, tk.END)
        entry_output_dir.insert(0, directorio)

# Crear los widgets
label_template_file = tk.Label(root, text="Archivo de Plantilla:")
label_template_file.grid(row=0, column=0, padx=10, pady=10)

entry_template_file = tk.Entry(root, width=50)
entry_template_file.grid(row=0, column=1, padx=10, pady=5)
button_template_file = tk.Button(root, text="Seleccionar", command=seleccionar_archivo)
button_template_file.grid(row=0, column=2, padx=10, pady=5)

label_output_dir = tk.Label(root, text="Directorio de Salida:")
label_output_dir.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
entry_output_dir = tk.Entry(root, width=50)
entry_output_dir.grid(row=1, column=1, padx=10, pady=5)
button_output_dir = tk.Button(root, text="Seleccionar", command=seleccionar_directorio)
button_output_dir.grid(row=1, column=2, padx=10, pady=5)

# Botón para ejecutar el proceso
button_ejecutar = tk.Button(root, text="Ejecutar", command=lambda: escribir_datos_en_archivo(entry_template_file.get(), entry_output_dir.get(), mapeo_campos_sauce, construcciones_data))
button_ejecutar.grid(row=2, column=1, padx=10, pady=20)

# Ejecutar la ventana principal
root.mainloop()
