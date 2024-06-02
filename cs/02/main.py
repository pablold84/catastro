import os
import shutil
import pandas as pd
from sqlalchemy import create_engine, inspect
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string
from PIL import Image as PILImage
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import logging

# Configurar logging
logging.basicConfig(filename='registro.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Función para configurar la conexión a la base de datos
def obtener_conexion():
    conn_params = {
        'host': 'cartofs.seresco.red',
        'dbname': 'SEGIPSA',
        'user': 'Segipsa',
        'password': 'Segipas24',
        'port': 5432
    }
    conn_str = f"postgresql+psycopg2://{conn_params['user']}:{conn_params['password']}@{conn_params['host']}:{conn_params['port']}/{conn_params['dbname']}"
    return create_engine(conn_str)

# Función para listar las tablas en la base de datos bajo un esquema dado
def listar_tablas(esquema):
    engine = obtener_conexion()
    inspector = inspect(engine)
    tablas = inspector.get_table_names(schema=esquema)
    return tablas

# Función para obtener REFCAT de la base de datos
def obtener_refcat(esquema):
    engine = obtener_conexion()
    tablas = listar_tablas(esquema)
    logging.info(f"Tablas disponibles en el esquema '{esquema}': {tablas}")
    
    # Asegúrate de que la tabla 'segipsa_placo' está en la lista de tablas
    if 'segipsa_placo' not in tablas:
        raise ValueError(f"La tabla 'segipsa_placo' no existe en el esquema '{esquema}'. Tablas disponibles: " + ", ".join(tablas))
    
    # Completa la consulta con esquema.tabla
    query = f'SELECT "REFCAT" FROM "{esquema}"."segipsa_placo" WHERE "accion" != \'CIER\''
    logging.info("Consulta ejecutada: " + query)
    df = pd.read_sql_query(query, engine)
    logging.info(f"Consulta exitosa. Resultado: {df}")
    return df['REFCAT'].tolist()

# Función para redimensionar y guardar las imágenes
def resize_image(image_path, output_path, width, height):
    with PILImage.open(image_path) as img:
        resized_img = img.resize((width, height), PILImage.Resampling.LANCZOS)
        resized_img.save(output_path)

# Función para añadir imágenes al archivo Excel
def add_images_to_excel(ws, png_path, jpg_path, desired_width, desired_height):
    resized_png_path = os.path.join(os.path.dirname(png_path), 'resized_' + os.path.basename(png_path))
    resized_jpg_path = os.path.join(os.path.dirname(jpg_path), 'resized_' + os.path.basename(jpg_path))
    
    if os.path.exists(png_path):
        resize_image(png_path, resized_png_path, desired_width, desired_height)
        img_png = OpenpyxlImage(resized_png_path)
        img_png.anchor = 'C5'
        ws.add_image(img_png)
    else:
        logging.warning(f"No se encontró la imagen PNG en la ruta: {png_path}")
    
    if os.path.exists(jpg_path):
        resize_image(jpg_path, resized_jpg_path, desired_width, desired_height)
        img_jpg = OpenpyxlImage(resized_jpg_path)
        img_jpg.anchor = 'R5'
        ws.add_image(img_jpg)
    else:
        logging.warning(f"No se encontró la imagen JPG en la ruta: {jpg_path}")

# Función para escribir datos iniciales en el Excel
def escribir_datos_iniciales(ws, esquema, refcat_value):
    engine = obtener_conexion()
    
    mapeo_campos = {
        "CARGO": "A",
        "ORD_CONS": "B",
        "ES": "C",
        "PLA": "D",
        "PUE": "E",
        "TIPOL": "F",
        "CAT_PREDO": "G",
        "DES": "H",
        "SUP_LOCAL": "I",
        "U_CONS": "J",
        "AP_CO_CO": "K",
        "ANY_ANTIG": "L"
    }

    start_row = 5
    query = f'SELECT * FROM "{esquema}"."DATOS_INICIALES" WHERE "REFCAT" = %(refcat)s'
    df_datos_iniciales = pd.read_sql_query(query, engine, params={'refcat': refcat_value})
    print("imprimo valor refcat_value ", refcat_value)
    print("imprimo valor consulta ", df_datos_iniciales)

    if df_datos_iniciales.empty:
        logging.error(f"No se encontraron registros con REFCAT = {refcat_value}")
        return False

    for idx, row in df_datos_iniciales.iterrows():
        for campo, col in mapeo_campos.items():
            if campo in row:
                valor_campo = row[campo]
                dest_col = column_index_from_string(col)
                dest_row = start_row + idx
                ws.cell(row=dest_row, column=dest_col, value=valor_campo)

    return True

# Función para escribir datos de SAUCE en el Excel
def escribir_datos_sauce(ws, archivo_csv):
    secciones = {"FINCAS": [], "EXPEDIENTE": [], "CONSTRUCCIONES": [], "UNIDADES CONSTRUCTIVAS": []}
    seccion_actual = None

    with open(archivo_csv, 'r') as file:
        for line in file:
            if line.strip() == "":
                seccion_actual = None
                continue

            if line.startswith("EXPEDIENTE"):
                seccion_actual = "EXPEDIENTE"
            elif line.startswith("FINCAS"):
                seccion_actual = "FINCAS"
            elif line.startswith("CONSTRUCCIONES"):
                seccion_actual = "CONSTRUCCIONES"
            elif line.startswith("UNIDADES CONSTRUCTIVAS"):
                seccion_actual = "UNIDADES CONSTRUCTIVAS"
            elif seccion_actual:
                secciones[seccion_actual].append(line.strip())

    construcciones_data = []
    for construccion in secciones["CONSTRUCCIONES"]:
        campos_construccion = construccion.split(";")
        construcciones_data.append([
            campos_construccion[29], campos_construccion[11], campos_construccion[13], 
            campos_construccion[14], campos_construccion[15], campos_construccion[19],
            campos_construccion[20], campos_construccion[18], campos_construccion[26],
            campos_construccion[16], "N/A", campos_construccion[23],
        campos_construccion[22]
        ])

    for unidad in secciones["UNIDADES CONSTRUCTIVAS"]:
        campos_unidad = unidad.split(";")
        for construccion in construcciones_data:
            if construccion[9] == campos_unidad[4]:
                construccion[10] = campos_unidad[16]
                break

    construcciones_data = construcciones_data[1:]
    mapeo_campos_sauce = {
        "CARGO": "O",
        "ORDEN": "P",
        "ESCALERA": "Q",
        "PLANTA": "R",
        "PUERTA": "S",
        "TIPOLOGIA": "T",
        "CATEG_PRED": "U",
        "DESTINO": "V",
        "SUPERFICIE_TOTAL": "W",
        "UNIDAD_CONST": "X",
        "COEF_CONSERVACION": "Y",
        "AA_ANTIGUEDAD": "Z",
        "AA_REFORMA": "AA"
    }

    start_row = 5
    for idx, construccion in enumerate(construcciones_data):
        for campo, col in mapeo_campos_sauce.items():
            valor_campo = construccion[list(mapeo_campos_sauce.keys()).index(campo)]
            ws[col + str(start_row + idx)] = valor_campo

# Función para comparar y resaltar diferencias
    # Obtener el rango de celdas a comparar (desde A5 hasta la última fila escrita en los datos iniciales)
def comparar_y_resaltar(ws):
    max_row = ws.max_row
    max_col_iniciales = ws.max_column
    max_col_sauce = max_col_iniciales + 14  # Ajuste para comenzar desde la columna "O" de SAUCE
    
    for row in range(5, max_row + 1):
        for col in range(1, max_col_iniciales + 1):
            cell_iniciales = ws.cell(row=row, column=col)
            cell_sauce = ws.cell(row=row, column=col + 14)  # Ajuste para alinear con la columna "O" de SAUCE
            
            # Obtener los valores de las celdas como texto
            valor_celda_iniciales = str(cell_iniciales.value)
            valor_celda_sauce = str(cell_sauce.value)
            
            # Comparar los valores de las celdas
            if valor_celda_iniciales != valor_celda_sauce:
                # Resaltar el texto en rojo en la parte de SAUCE
                cell_sauce.font = Font(color="FF0000")  # Rojo

# Función para procesar cada carpeta en el directorio de origen
def process_folders(window, output_dir, template_file, origin_dir, esquema, progress_label, progress_bar):
    refcat_list = obtener_refcat(esquema)
    total_folders = len(refcat_list)

    for index, refcat_value in enumerate(refcat_list):
        refcat_value = refcat_value.strip()  # Asegúrate de que no haya espacios en blanco
        folder_path = os.path.join(origin_dir, refcat_value)  # Modificado para usar la ruta de origen
        excel_path = os.path.join(output_dir, f"{refcat_value}_FichaResumen.xlsx")
        png_path = os.path.join(folder_path, f"{refcat_value}.png")
        jpg_path = os.path.join(folder_path, f"{refcat_value}.jpg")
        csv_path = os.path.join(folder_path, f"{refcat_value}.csv")
        
        logging.info(f"Refcat value: {refcat_value}")
        
        # Actualizar la barra de progreso
        progress_bar['value'] = (index + 1) / total_folders * 100
        progress_label.config(text=f"Procesando carpeta {index + 1} de {total_folders}")
        window.update_idletasks()
        
        # Verificar si los archivos PNG, JPG y CSV existen en la carpeta actual
        if not all(os.path.exists(path) for path in [png_path, jpg_path, csv_path]):
            logging.warning(f"Archivos faltantes para REFCAT {refcat_value}")
            missing_files = [path for path in [png_path, jpg_path, csv_path] if not os.path.exists(path)]
            for missing_file in missing_files:
                logging.warning(f"El archivo {missing_file} no se encuentra en la carpeta {folder_path}")
            continue
        
        # Verificar si la carpeta de salida existe
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)  # Asegurarse de que la carpeta de salida existe
            logging.info(f"Carpeta de salida creada en: {output_dir}")
        
        # Copiar la plantilla al directorio de salida
        shutil.copy(template_file, excel_path)
        logging.info(f"Plantilla copiada a: {excel_path}")
        
        # Cargar el libro de trabajo y las hojas necesarias
        wb = openpyxl.load_workbook(excel_path)
        ws_iniciales = wb['SAUCE']
        ws_sauce = wb['SAUCE']
        ws_croquis = wb['CROQUIS']
        
        # Escribir datos iniciales en el libro de trabajo
        if not escribir_datos_iniciales(ws_iniciales, esquema, refcat_value):
            continue
        
        # Escribir datos de SAUCE en el libro de trabajo
        escribir_datos_sauce(ws_sauce, csv_path)
        
        # Agregar imágenes al libro de trabajo
        add_images_to_excel(ws_croquis, png_path, jpg_path, 700, 700)
        
        # Comparar y resaltar diferencias en los datos de SAUCE
        comparar_y_resaltar(ws_sauce)
        
        # Guardar el libro de trabajo modificado
        wb.save(excel_path)
        logging.info(f"Libro de trabajo guardado en: {excel_path}")

# Función para abrir el explorador de archivos
def seleccionar_directorio(titulo):
    root = tk.Tk()
    root.withdraw()
    directorio = filedialog.askdirectory(title=titulo)
    return directorio

def seleccionar_archivo(titulo):
    root = tk.Tk()
    root.withdraw()
    archivo = filedialog.askopenfilename(title=titulo, filetypes=[('Excel Files', '*.xlsx')])
    return archivo

# Función para la interfaz gráfica
def interfaz_grafica():
    window = tk.Tk()
    window.title("Procesar Archivos Masivamente")

    ttk.Label(window, text="Esquema:").grid(column=0, row=0)
    schema_entry = ttk.Entry(window, width=50)
    schema_entry.grid(column=1, row=0)
    
    ttk.Label(window, text="Directorio de Salida:").grid(column=0, row=1)
    output_dir_entry = ttk.Entry(window, width=50)
    output_dir_entry.grid(column=1, row=1)
    ttk.Button(window, text="Seleccionar", command=lambda: output_dir_entry.insert(0, seleccionar_directorio("Seleccionar Directorio de Salida"))).grid(column=2, row=1)
    
    ttk.Label(window, text="Archivo de Plantilla:").grid(column=0, row=2)
    template_file_entry = ttk.Entry(window, width=50)
    template_file_entry.grid(column=1, row=2)
    ttk.Button(window, text="Seleccionar", command=lambda: template_file_entry.insert(0, seleccionar_archivo("Seleccionar Archivo de Plantilla"))).grid(column=2, row=2)
    
    ttk.Label(window, text="Directorio de Origen de Datos:").grid(column=0, row=3)
    origin_dir_entry = ttk.Entry(window, width=50)
    origin_dir_entry.grid(column=1, row=3)
    ttk.Button(window, text="Seleccionar", command=lambda: origin_dir_entry.insert(0, seleccionar_directorio("Seleccionar Directorio de Origen de Datos"))).grid(column=2, row=3)
    
    progress_label = ttk.Label(window, text="")
    progress_label.grid(column=0, row=4, columnspan=3)

    progress_bar = ttk.Progressbar(window, orient=tk.HORIZONTAL, length=300, mode='determinate')
    progress_bar.grid(column=0, row=5, columnspan=3)
        
    def ejecutar_proceso():
        esquema = schema_entry.get()
        output_dir = output_dir_entry.get()
        template_file = template_file_entry.get()
        origin_dir = origin_dir_entry.get()
        
        if not all([esquema, output_dir, template_file, origin_dir]):
            messagebox.showerror("Error", "Todos los campos son obligatorios.")
            return
        
        process_folders(window, output_dir, template_file, origin_dir, esquema, progress_label, progress_bar)
        messagebox.showinfo("Éxito", "Proceso completado.")

    ttk.Button(window, text="Ejecutar", command=ejecutar_proceso).grid(column=0, row=6, columnspan=3)
    
    window.mainloop()

# Ejecutar la interfaz gráfica
if __name__ == "__main__":
    interfaz_grafica()
