import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox

# Función para configurar la conexión a la base de datos
def obtener_conexion():
    conn_params = {
        'host': 'cartofs.seresco.red',
        'dbname': 'SEGIPSA',
        'user': 'Segipsa',
        'password': 'Segipas24',
        'port': 5432
    }
    conn_str = f"postgresql+psycopg2://{conn_params['user']}:{conn_params['password']}@{conn_params['host']}:{conn_params['port']}/{conn_params['dbname']}"
    return create_engine(conn_str)

# Función para ejecutar el proceso
def ejecutar_proceso(output_dir, template_file, schema):
    # Conectar a la base de datos
    engine = obtener_conexion()
    
    # Obtener todos los registros de la tabla DATOS_INICIALES filtrando por REFCAT
    refcat_value = '001501800UN59G'  # Valor específico de REFCAT
    df_datos_iniciales = pd.read_sql_query(f'SELECT * FROM "{schema}"."DATOS_INICIALES" WHERE "REFCAT" = %s', engine, params=(refcat_value,))
    
    # Verificar si se encontraron registros
    if df_datos_iniciales.empty:
        messagebox.showerror("Error", f"No se encontraron registros con REFCAT = {refcat_value}")
        return

    # Campos y celdas de destino para DATOS_INICIALES, ajustados para comenzar en A5
    mapeo_campos_datos_iniciales = {
        "CARGO": "A",
        "ORD_CONS": "B",
        "ES": "C",
        "PLA": "D",
        "PUE": "E",
        "TIPOL": "F",
        "CAT_PREDO": "G",
        "DES": "H",
        "SUP_LOCAL": "I",
        "U_CONS": "J",
        "AP_CO_CO": "K",
        "ANY_ANTIG": "L"
    }

    # Definir mapeo_campos_sauce
    mapeo_campos_sauce = {
        "CARGO": "O",
        "ORD_CONS": "P",
        "ES": "Q",
        "PLA": "R",
        "PUE": "S",
        "TIPOL": "T",
        "CAT_PREDO": "U",
        "DES": "V",
        "SUP_LOCAL": "W",
        "U_CONS": "X",
        "AP_CO_CO": "Y",
        "ANY_ANTIG": "Z"
    }

    # Crear el directorio de salida si no existe
    os.makedirs(output_dir, exist_ok=True)
    
    # Usar el campo REFCAT como identificador en el nombre del archivo
    file_name = f"FICHA_RESUMEN_PLACO23_{refcat_value}.xlsx"
    file_path = os.path.join(output_dir, file_name)

    # Copiar el archivo de plantilla
    shutil.copyfile(template_file, file_path)

    # Cargar el archivo copiado
    book = load_workbook(file_path)
    ws = book["SAUCE"]

    # Escribir los datos de DATOS_INICIALES en el Excel
    start_row_datos_iniciales = 5
    for idx, row in df_datos_iniciales.iterrows():
        for campo, col in mapeo_campos_datos_iniciales.items():
            if campo in row:
                valor_campo = row[campo]
                dest_col = column_index_from_string(col)
                dest_row = start_row_datos_iniciales + idx
                ws.cell(row=dest_row, column=dest_col, value=valor_campo)

    # Carga de la información "SITUACIÓN FINAL datos sauce"
    archivo_csv = "modelo6/origen/ficheros/sauce/33_217210_24.csv"
    secciones = {"FINCAS": [], "EXPEDIENTE": [], "CONSTRUCCIONES":[], "UNIDADES CONSTRUCTIVAS":[]}
    seccion_actual = None

    with open(archivo_csv, 'r') as file:
        for line in file:
            if line.strip() == "":
                seccion_actual = None
                continue
            
            if line.startswith("EXPEDIENTE"):
                seccion_actual = "EXPEDIENTE"
                continue
            elif line.startswith("FINCAS"):
                seccion_actual = "FINCAS"
                continue
            elif line.startswith("CONSTRUCCIONES"):
                seccion_actual = "CONSTRUCCIONES"
                continue
            elif line.startswith("UNIDADES CONSTRUCTIVAS"):
                seccion_actual = "UNIDADES CONSTRUCTIVAS"
                continue
            
            if seccion_actual in secciones and line.strip() != "":
                secciones[seccion_actual].append(line.strip())

    construcciones_data = []
    for construccion in secciones["CONSTRUCCIONES"]:
        campos_construccion = construccion.split(";")
        cargo = campos_construccion[29]
        orden = campos_construccion[11]
        pcatastral1 = campos_construccion[9]
        pcatastral2 = campos_construccion[10]
        escalera = campos_construccion[13]
        planta = campos_construccion[14]
        puerta = campos_construccion[15]
        tipologia = campos_construccion[19]
        categ_pred = campos_construccion[20]
        destino = campos_construccion[18]
        superficie_total = campos_construccion[26]
        unidad_const_dest = campos_construccion[16]
        aa_antiguedad = campos_construccion[23]
        aa_reforma = campos_construccion[22]

        coef_conservacion = "N/A"
        for unidad in secciones["UNIDADES CONSTRUCTIVAS"]:
            campos_unidad = unidad.split(";")
            if pcatastral1 == campos_unidad[2] and pcatastral2 == campos_unidad[3] and unidad_const_dest == campos_unidad[4]:
                coef_conservacion = campos_unidad[16]
                break

    construcciones_data = []
    for construccion in secciones["CONSTRUCCIONES"]:
        campos_construccion = construccion.split(";")
        construccion_dict = {
            "CARGO": campos_construccion[29],
            "ORD_CONS": campos_construccion[11],
            "ES": campos_construccion[13],
            "PLA": campos_construccion[14],
            "PUE": campos_construccion[15],
            "TIPOL": campos_construccion[19],
            "CAT_PREDO": campos_construccion[20],
            "DES": campos_construccion[18],
            "SUP_LOCAL": campos_construccion[26],
            "U_CONS": campos_construccion[16],
            "AP_CO_CO": campos_construccion[22],
            "ANY_ANTIG": campos_construccion[23]
        }
        construcciones_data.append(construccion_dict)


    # Escribir los datos de "SITUACIÓN FINAL datos sauce" en el Excel
    start_row_sauce = start_row_datos_iniciales
    for idx, construccion in enumerate(construcciones_data):
        for campo, col in mapeo_campos_sauce.items():
            if campo in construccion:
                valor_campo = construccion[mapeo_campos_sauce.get(campo, -1)]  # Usamos .get() para manejar casos donde campo no está presente en mapeo_campos_sauce
                dest_col = column_index_from_string(col)
                dest_row = start_row_sauce + idx
                ws.cell(row=dest_row, column=dest_col, value=valor_campo)

    # Guardar el libro de Excel
    book.save(file_path)
    
    messagebox.showinfo("Éxito", "Datos escritos exitosamente en el archivo Excel generado")

# Función para seleccionar el directorio de salida
def seleccionar_directorio():
    directorio = filedialog.askdirectory()
    if directorio:
        entry_output_dir.delete(0, tk.END)
        entry_output_dir.insert(0, directorio)

# Función para seleccionar el archivo de plantilla
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if archivo:
        entry_template_file.delete(0, tk.END)
        entry_template_file.insert(0, archivo)

# Crear la ventana principal
root = tk.Tk()
root.title("Generador de Modelo 6")

# Crear los widgets
label_output_dir = tk.Label(root, text="Directorio de Salida:")
label_output_dir.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
entry_output_dir = tk.Entry(root, width=50)
entry_output_dir.grid(row=0, column=1, padx=10, pady=5)
button_output_dir = tk.Button(root, text="Seleccionar", command=seleccionar_directorio)
button_output_dir.grid(row=0, column=2, padx=10, pady=5)

label_template_file = tk.Label(root, text="Archivo de Plantilla:")
label_template_file.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
entry_template_file = tk.Entry(root, width=50)
entry_template_file.grid(row=1, column=1, padx=10, pady=5)
button_template_file = tk.Button(root, text="Seleccionar", command=seleccionar_archivo)
button_template_file.grid(row=1, column=2, padx=10, pady=5)

label_schema = tk.Label(root, text="Esquema:")
label_schema.grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
entry_schema = tk.Entry(root, width=50)
entry_schema.grid(row=2, column=1, padx=10, pady=5)

button_ejecutar = tk.Button(root, text="Ejecutar", command=lambda: ejecutar_proceso(entry_output_dir.get(), entry_template_file.get(), entry_schema.get()))
button_ejecutar.grid(row=3, column=1, padx=10, pady=20)

# Ejecutar la ventana principal
root.mainloop()

