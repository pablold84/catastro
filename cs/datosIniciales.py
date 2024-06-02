import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox

# Función para configurar la conexión a la base de datos
def obtener_conexion():
    conn_params = {
        'host': 'cartofs.seresco.red',
        'dbname': 'SEGIPSA',
        'user': 'Segipsa',
        'password': 'Segipas24',
        'port': 5432
    }
    conn_str = f"postgresql+psycopg2://{conn_params['user']}:{conn_params['password']}@{conn_params['host']}:{conn_params['port']}/{conn_params['dbname']}"
    return create_engine(conn_str)

# Función para ejecutar el proceso
def ejecutar_proceso(output_dir, template_file, schema):
    # Conectar a la base de datos
    engine = obtener_conexion()
    
    # Obtener todos los registros de la tabla DATOS_INICIALES filtrando por REFCAT
    refcat_value = '001501800UN59G'  # Valor específico de REFCAT
    df_datos_iniciales = pd.read_sql_query(f'SELECT * FROM "{schema}"."DATOS_INICIALES" WHERE "REFCAT" = %s', engine, params=(refcat_value,))
    
    # Imprimir los nombres de las columnas para verificar
    print("Columnas en df_datos_iniciales:", df_datos_iniciales.columns)
    
    # Verificar si se encontraron registros
    if df_datos_iniciales.empty:
        messagebox.showerror("Error", f"No se encontraron registros con REFCAT = {refcat_value}")
        return

    # Campos y celdas de destino para DATOS_INICIALES, ajustados para comenzar en A5
    mapeo_campos = {
        "CARGO": "A",
        "ORD_CONS": "B",
        "ES": "C",
        "PLA": "D",
        "PUE": "E",
        "TIPOL": "F",
        "CAT_PREDO": "G",
        "DES": "H",
        "SUP_LOCAL": "I",
        "U_CONS": "J",
        "AP_CO_CO": "K",
        "ANY_ANTIG": "L"
    }


    # Crear el directorio de salida si no existe
    os.makedirs(output_dir, exist_ok=True)
    
    # Usar el campo REFCAT como identificador en el nombre del archivo
    file_name = f"FICHA_RESUMEN_PLACO23_{refcat_value}.xlsx"
    file_path = os.path.join(output_dir, file_name)

    # Copiar el archivo de plantilla
    shutil.copyfile(template_file, file_path)

    # Cargar el archivo copiado
    book = load_workbook(file_path)
    ws = book["SAUCE"]

    # Escribir los datos de DATOS_INICIALES en el Excel
    start_row = 5
    for idx, row in df_datos_iniciales.iterrows():
        for campo, col in mapeo_campos.items():
            if campo in row:
                valor_campo = row[campo]
                dest_col = column_index_from_string(col)
                dest_row = start_row + idx
                ws.cell(row=dest_row, column=dest_col, value=valor_campo)

    # Guardar el libro de Excel
    book.save(file_path)
    
    messagebox.showinfo("Éxito", "Datos escritos exitosamente en el archivo Excel generado")

# Función para seleccionar el directorio de salida
def seleccionar_directorio():
    directorio = filedialog.askdirectory()
    if directorio:
        entry_output_dir.delete(0, tk.END)
        entry_output_dir.insert(0, directorio)

# Función para seleccionar el archivo de plantilla
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if archivo:
        entry_template_file.delete(0, tk.END)
        entry_template_file.insert(0, archivo)

# Crear la ventana principal
root = tk.Tk()
root.title("Generador de Modelo 6")

# Crear los widgets
label_output_dir = tk.Label(root, text="Directorio de Salida:")
label_output_dir.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
entry_output_dir = tk.Entry(root, width=50)
entry_output_dir.grid(row=0, column=1, padx=10, pady=5)
button_output_dir = tk.Button(root, text="Seleccionar", command=seleccionar_directorio)
button_output_dir.grid(row=0, column=2, padx=10, pady=5)

label_template_file = tk.Label(root, text="Archivo de Plantilla:")
label_template_file.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
entry_template_file = tk.Entry(root, width=50)
entry_template_file.grid(row=1, column=1, padx=10, pady=5)
button_template_file = tk.Button(root, text="Seleccionar", command=seleccionar_archivo)
button_template_file.grid(row=1, column=2, padx=10, pady=5)

label_schema = tk.Label(root, text="Esquema:")
label_schema.grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
entry_schema = tk.Entry(root, width=50)
entry_schema.grid(row=2, column=1, padx=10, pady=5)

button_ejecutar = tk.Button(root, text="Ejecutar", command=lambda: ejecutar_proceso(entry_output_dir.get(), entry_template_file.get(), entry_schema.get()))
button_ejecutar.grid(row=3, column=1, padx=10, pady=20)

# Ejecutar la ventana principal
root.mainloop()
