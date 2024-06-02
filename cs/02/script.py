import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import os
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox

# Función para configurar la conexión a la base de datos
def obtener_conexion():
    conn_params = {
        'host': 'cartofs.seresco.red',
        'dbname': 'SEGIPSA',
        'user': 'Segipsa',
        'password': 'Segipas24',
        'port': 5432
    }
    conn_str = f"postgresql+psycopg2://{conn_params['user']}:{conn_params['password']}@{conn_params['host']}:{conn_params['port']}/{conn_params['dbname']}"
    return create_engine(conn_str)

# Función para ejecutar el proceso
def ejecutar_proceso(output_dir, template_file):
    # Conectar a la base de datos
    engine = obtener_conexion()
    
    # Obtener todos los registros de la base de datos
    df = pd.read_sql_query("SELECT * FROM cabrales.segipsa_placo where exp='217210' or exp='217431'", engine)
    
    # Definir consultas SQL y campos correspondientes
    consultas_campos = {
        "campo_1": {
            "consulta": "SELECT del FROM cabrales.segipsa_placo",
            "campo": "del",
            "texto_agregar": "GERENCIA-MUNICIPIO: ",
            "celda_destino": "A7"
        },
        "campo_2": {
            "consulta": "SELECT exp, control, anio FROM cabrales.segipsa_placo",
            "campo": "expediente",
            "texto_agregar": "Nº EXPEDIENTE: ",
            "celda_destino": "P7"
        },
        "campo_3": {
            "consulta": "SELECT fecha_proyecto FROM cabrales.segipsa_placo",
            "campo": "fecha_proyecto",
            "texto_agregar": "FECHA: ",
            "celda_destino": "H21"  # Modifica la celda de destino según corresponda
        }
    }
    
    # Crear el directorio de salida si no existe
    os.makedirs(output_dir, exist_ok=True)
    
    # Iterar sobre cada registro y generar un archivo Excel
    for idx, row in df.iterrows():
        # Nombre del archivo usando el campo 'exp'
        exp_value = row['exp']
        file_name = f"FICHA_RESUMEN_PLACO23_{exp_value}.xlsx"
        file_path = os.path.join(output_dir, file_name)

        # Copiar el archivo de plantilla
        shutil.copyfile(template_file, file_path)

        # Cargar el archivo copiado
        book = load_workbook(file_path)
        ws = book["FICHA RESUMEN PLACO"]

        # Escribir los datos en el Excel
        for nombre_campo, detalles in consultas_campos.items():
            texto_agregar = detalles["texto_agregar"]
            celda_destino = detalles["celda_destino"]
            consulta_sql = detalles["consulta"]

            # Leer los datos de la base de datos PostgreSQL en un DataFrame de pandas
            df_consulta = pd.read_sql_query(consulta_sql, engine)

            # Procesar el campo_2
            if nombre_campo == "campo_2":
                expediente = str(row["exp"]) + "." + str(row["control"]) + "/" + str(row["anio"])
                texto_agregar += expediente
                initial_col, dest_row = column_index_from_string(celda_destino[0]), int(celda_destino[1:])
                ws.cell(row=dest_row, column=initial_col, value=texto_agregar)
            
            # Procesar el campo_3
            elif nombre_campo == "campo_3":
                # Obtener la fecha de proyecto desde el DataFrame
                fecha_proyecto = row['fecha_proyecto']

                # Verificar si la fecha de proyecto tiene valor
                if pd.notnull(fecha_proyecto) and fecha_proyecto != "":
                    # Convertir la fecha a cadena si es necesario
                    fecha_proyecto_str = str(fecha_proyecto)

                    # Activar el checkbox en la celda G20
                    ws['G20'].value = "x"
                    
                    # Escribir la fecha en la celda H21
                    ws['H21'].value = f"{texto_agregar}{fecha_proyecto_str}"
           
            
            # Procesar otros campos
            else:
                valor_campo = row[detalles["campo"]]
                dest_col, dest_row = column_index_from_string(celda_destino[0]), int(celda_destino[1:])
                ws.cell(row=dest_row, column=dest_col, value=texto_agregar + str(valor_campo))

        # Guardar el libro de Excel
        book.save(file_path)
    
    messagebox.showinfo("Éxito", "Datos escritos exitosamente en los archivos Excel generados")

# Función para seleccionar el directorio de salida
def seleccionar_directorio():
    directorio = filedialog.askdirectory()
    if directorio:
        entry_output_dir.delete(0, tk.END)
        entry_output_dir.insert(0, directorio)

# Función para seleccionar el archivo de plantilla
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
    if archivo:
        entry_template_file.delete(0, tk.END)
        entry_template_file.insert(0, archivo)

# Crear la ventana principal
root = tk.Tk()
root.title("Generador de Modelo 6")

# Crear los widgets
label_output_dir = tk.Label(root, text="Directorio de Salida:")
label_output_dir.grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
entry_output_dir = tk.Entry(root, width=50)
entry_output_dir.grid(row=0, column=1, padx=10, pady=5)
button_output_dir = tk.Button(root, text="Seleccionar", command=seleccionar_directorio)
button_output_dir.grid(row=0, column=2, padx=10, pady=5)

label_template_file = tk.Label(root, text="Archivo de Plantilla:")
label_template_file.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
entry_template_file = tk.Entry(root, width=50)
entry_template_file.grid(row=1, column=1, padx=10, pady=5)
button_template_file = tk.Button(root, text="Seleccionar", command=seleccionar_archivo)
button_template_file.grid(row=1, column=2, padx=10, pady=5)

button_ejecutar = tk.Button(root, text="Ejecutar", command=lambda: ejecutar_proceso(entry_output_dir.get(), entry_template_file.get()))
button_ejecutar.grid(row=2, column=1, padx=10, pady=20)

# Ejecutar la ventana principal
root.mainloop()
